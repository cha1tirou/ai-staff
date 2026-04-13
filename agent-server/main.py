import asyncio
import os
import json
import smtplib
import tempfile
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from fastapi import FastAPI, HTTPException, Header
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
from claude_agent_sdk import query, ClaudeAgentOptions, AgentDefinition

app = FastAPI()

INTERNAL_SECRET = os.environ.get("INTERNAL_SECRET", "")
GMAIL_USER = os.environ.get("GMAIL_USER", "")
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "")

# リクエスト型
class SubAgent(BaseModel):
    name: str
    role: str
    tasks: list[str]
    custom_tasks: str = ""
    report_email: str = ""

class RunRequest(BaseModel):
    message: str
    biz_name: str
    sub_agents: list[SubAgent]
    report_email: str = ""  # 送信先（エージェントの report_email）

# ── メール送信ユーティリティ ──
def send_email_with_attachment(to: str, subject: str, body: str, filepath: str, filename: str):
    """ファイルをメール添付して送信する"""
    if not GMAIL_USER or not GMAIL_APP_PASSWORD:
        return False, "GMAIL_USER / GMAIL_APP_PASSWORD が未設定です"

    msg = MIMEMultipart()
    msg["From"] = GMAIL_USER
    msg["To"] = to
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain", "utf-8"))

    with open(filepath, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
        msg.attach(part)

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(GMAIL_USER, GMAIL_APP_PASSWORD)
            smtp.sendmail(GMAIL_USER, to, msg.as_string())
        return True, "送信成功"
    except Exception as e:
        return False, str(e)

# ── カスタムツール定義 ──
CUSTOM_TOOLS = [
    {
        "name": "generate_excel",
        "description": "Excelファイルを生成してメールで送信する。表形式のデータレポートや集計表に使う。",
        "input_schema": {
            "type": "object",
            "properties": {
                "filename": {"type": "string", "description": "ファイル名（拡張子なし）"},
                "sheet_name": {"type": "string", "description": "シート名"},
                "headers": {"type": "array", "items": {"type": "string"}, "description": "列ヘッダーのリスト"},
                "rows": {"type": "array", "items": {"type": "array"}, "description": "データ行のリスト"},
                "email_subject": {"type": "string", "description": "メールの件名"},
                "email_body": {"type": "string", "description": "メール本文"},
                "to_email": {"type": "string", "description": "送信先メールアドレス"}
            },
            "required": ["filename", "headers", "rows", "email_subject", "email_body", "to_email"]
        }
    },
    {
        "name": "generate_pdf",
        "description": "PDFレポートを生成してメールで送信する。調査報告書・提案書・議事録などに使う。",
        "input_schema": {
            "type": "object",
            "properties": {
                "filename": {"type": "string", "description": "ファイル名（拡張子なし）"},
                "title": {"type": "string", "description": "PDFのタイトル"},
                "sections": {
                    "type": "array",
                    "description": "セクションのリスト",
                    "items": {
                        "type": "object",
                        "properties": {
                            "heading": {"type": "string"},
                            "content": {"type": "string"}
                        }
                    }
                },
                "email_subject": {"type": "string", "description": "メールの件名"},
                "email_body": {"type": "string", "description": "メール本文"},
                "to_email": {"type": "string", "description": "送信先メールアドレス"}
            },
            "required": ["filename", "title", "sections", "email_subject", "email_body", "to_email"]
        }
    }
]

def handle_generate_excel(input_data: dict, report_email: str) -> str:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = input_data.get("sheet_name", "Sheet1")

        # ヘッダー行
        headers = input_data["headers"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # データ行
        for row_idx, row in enumerate(input_data["rows"], 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 列幅を自動調整
        for col in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # 一時ファイルに保存
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmppath = tmp.name
        wb.save(tmppath)

        filename = f"{input_data['filename']}.xlsx"
        to = input_data.get("to_email") or report_email
        ok, msg = send_email_with_attachment(
            to, input_data["email_subject"], input_data["email_body"], tmppath, filename
        )
        os.unlink(tmppath)

        if ok:
            return f"✅ Excelファイル「{filename}」を {to} に送信しました。"
        else:
            return f"❌ メール送信失敗: {msg}"
    except Exception as e:
        return f"❌ Excel生成エラー: {str(e)}"

def handle_generate_pdf(input_data: dict, report_email: str) -> str:
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
        from reportlab.lib.enums import TA_LEFT
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont

        pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))

        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmppath = tmp.name

        doc = SimpleDocTemplate(tmppath, pagesize=A4,
                                rightMargin=20*mm, leftMargin=20*mm,
                                topMargin=20*mm, bottomMargin=20*mm)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("Title", fontName="HeiseiKakuGo-W5",
                                     fontSize=18, leading=24, spaceAfter=12)
        heading_style = ParagraphStyle("Heading", fontName="HeiseiKakuGo-W5",
                                       fontSize=13, leading=18, spaceAfter=6,
                                       textColor="#1e40af")
        body_style = ParagraphStyle("Body", fontName="HeiseiKakuGo-W5",
                                    fontSize=10, leading=16, spaceAfter=8)

        story = []
        story.append(Paragraph(input_data["title"], title_style))
        story.append(Spacer(1, 8*mm))

        for section in input_data.get("sections", []):
            if section.get("heading"):
                story.append(Paragraph(section["heading"], heading_style))
            if section.get("content"):
                for line in section["content"].split("\n"):
                    if line.strip():
                        story.append(Paragraph(line, body_style))
            story.append(Spacer(1, 4*mm))

        doc.build(story)

        filename = f"{input_data['filename']}.pdf"
        to = input_data.get("to_email") or report_email
        ok, msg = send_email_with_attachment(
            to, input_data["email_subject"], input_data["email_body"], tmppath, filename
        )
        os.unlink(tmppath)

        if ok:
            return f"✅ PDFファイル「{filename}」を {to} に送信しました。"
        else:
            return f"❌ メール送信失敗: {msg}"
    except Exception as e:
        return f"❌ PDF生成エラー: {str(e)}"

def build_orchestrator_prompt(biz_name: str, sub_agents: list[SubAgent], report_email: str) -> str:
    staff_list = "\n".join([
        f"- {a.name}（{a.role}）: {', '.join(a.tasks)}"
        for a in sub_agents
    ])
    return f"""あなたは{biz_name}のAIスタッフを束ねるマネージャーです。

【管理下のスタッフ】
{staff_list}

【使えるツール】
- generate_excel: Excelファイルを生成してメール送信
- generate_pdf: PDFファイルを生成してメール送信
- Web検索（Bash/ウェブ）: 最新情報の調査
- Agent: サブエージェントへの委譲

【デフォルト送信先】
{report_email}

【あなたの役割】
- ユーザーの指示を受け取り、最適なスタッフに振り分ける
- ファイル生成が必要なタスクはgenerate_excel/generate_pdfツールを使う
- 必ず日本語で返答する
- 実行した内容を透明性高くユーザーに伝える"""

def build_sub_agent_definition(agent: SubAgent) -> AgentDefinition:
    task_list = "\n- ".join(filter(None, agent.tasks + [agent.custom_tasks]))
    return AgentDefinition(
        description=f"{agent.role}の専門エージェント。{', '.join(agent.tasks)}に関するタスクを担当。",
        prompt=f"""あなたは{agent.name}として働くAIスタッフです。

【担当業務】
- {task_list}

【ルール】
- 担当業務に関する質問・タスクに専門的に答える
- ファイル生成が必要な場合はgenerate_excel/generate_pdfツールを使う
- 簡潔かつ具体的に回答する
- 日本語で答える""",
        model="sonnet",
        tools=["WebSearch", "Agent"],
    )

@app.post("/run")
async def run_agent(
    req: RunRequest,
    x_internal_secret: Optional[str] = Header(None)
):
    if INTERNAL_SECRET and x_internal_secret != INTERNAL_SECRET:
        raise HTTPException(status_code=401, detail="Unauthorized")

    if not req.sub_agents:
        raise HTTPException(status_code=400, detail="sub_agents is required")

    report_email = req.report_email or (req.sub_agents[0].report_email if req.sub_agents else "")
    agents = {agent.name: build_sub_agent_definition(agent) for agent in req.sub_agents}
    orch_prompt = build_orchestrator_prompt(req.biz_name, req.sub_agents, report_email)

    async def stream():
        try:
            async for message in query(
                prompt=req.message,
                options=ClaudeAgentOptions(
                    system_prompt=orch_prompt,
                    allowed_tools=["WebSearch", "Agent"],
                    agents=agents,
                    permission_mode="bypassPermissions",
                    custom_tools=CUSTOM_TOOLS,
                ),
            ):
                # カスタムツール呼び出しを処理
                if hasattr(message, "message"):
                    for block in getattr(message.message, "content", []):
                        btype = getattr(block, "type", None)
                        if btype == "tool_use":
                            tool_name = getattr(block, "name", "")
                            tool_input = getattr(block, "input", {})
                            if tool_name == "generate_excel":
                                result = handle_generate_excel(tool_input, report_email)
                                yield f"data: {result}\n\n"
                            elif tool_name == "generate_pdf":
                                result = handle_generate_pdf(tool_input, report_email)
                                yield f"data: {result}\n\n"
                        elif btype == "text" and block.text:
                            yield f"data: {block.text}\n\n"

                if hasattr(message, "result") and message.result:
                    yield f"data: {message.result}\n\n"

        except Exception as e:
            yield f"data: [ERROR] {str(e)}\n\n"

    return StreamingResponse(stream(), media_type="text/event-stream")

@app.get("/health")
def health():
    return {"status": "ok"}
